#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo para exportação de dados contábeis para Excel.

Gera arquivo Excel com múltiplas abas:
- Plano de Contas
- Balanço Patrimonial
- DRE (Demonstração do Resultado do Exercício)
- Movimentação do Período
"""
from datetime import date, timedelta
from pathlib import Path
from typing import Optional, Dict, List
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pyaccount.db_client import ContabilDBClient
from pyaccount.classificacao import AccountClassifier, CLASSIFICACAO_M1
from pyaccount.account_mapper import AccountMapper
from pyaccount.utils import normalizar_nome


class ExcelExporter:
    """
    Exportador de dados contábeis para Excel.
    
    Esta classe encapsula a lógica necessária para:
    - Buscar plano de contas com saldos finais
    - Gerar Balanço Patrimonial estruturado
    - Gerar DRE com cálculos
    - Exportar movimentações do período
    - Formatar e exportar tudo para Excel
    """
    
    def __init__(
        self,
        db_client: ContabilDBClient,
        empresa: int,
        inicio: date,
        fim: date,
        classificacao_customizada: Optional[Dict[str, str]] = None,
        desconsiderar_zeramento: bool = True,
    ):
        """
        Inicializa o exportador Excel.
        
        Args:
            db_client: Cliente de banco de dados
            empresa: Código da empresa
            inicio: Data inicial do período
            fim: Data final do período
            classificacao_customizada: Dicionário opcional com mapeamento customizado
                                      de prefixos CLAS_CTA para categorias Beancount
            desconsiderar_zeramento: Se True, exclui lançamentos com orig_lan = 2 (Zeramento)
        """
        self.db_client = db_client
        self.empresa = empresa
        self.inicio = inicio
        self.fim = fim
        self.classificacao_customizada = classificacao_customizada
        self.desconsiderar_zeramento = desconsiderar_zeramento
        
        # Mapeador de contas (classe base compartilhada)
        self.account_mapper = AccountMapper(classificacao_customizada)
        
        # DataFrames internos
        self.df_pc: Optional[pd.DataFrame] = None
        self.df_saldos_finais: Optional[pd.DataFrame] = None
        self.df_saldos_iniciais: Optional[pd.DataFrame] = None
        self.df_movimentacoes: Optional[pd.DataFrame] = None
        self.df_lancamentos: Optional[pd.DataFrame] = None
        self.mapa_codi_to_bc: Dict[str, str] = {}
    
    def classificar_beancount(self, clas_cta: str, tipo_cta: Optional[str] = None) -> str:
        """Mapeia CLAS_CTA -> grupo Beancount."""
        return self.account_mapper.classificar_beancount(clas_cta, tipo_cta)
    
    def buscar_plano_contas_com_saldos(self) -> pd.DataFrame:
        """
        Busca plano de contas do banco de dados com saldos finais.
        
        Returns:
            DataFrame com plano de contas e saldos finais
        """
        # Busca plano de contas
        df_pc = self.db_client.buscar_plano_contas(self.empresa)
        
        if df_pc.empty:
            raise RuntimeError("Plano de contas vazio para a empresa informada.")
        
        # Processa plano de contas usando AccountMapper
        df_pc = self.account_mapper.processar_plano_contas(df_pc, filtrar_ativas=False)
        
        # Cria mapa para lookup
        mapas = self.account_mapper.criar_mapas(df_pc)
        self.mapa_codi_to_bc = mapas["codi_to_bc"]
        
        # Busca saldos finais
        df_saldos = self.db_client.buscar_saldos(self.empresa, self.fim)
        
        # Mescla saldos ao plano de contas
        if not df_saldos.empty:
            df_saldos["conta"] = df_saldos["conta"].astype(str)
            df_pc["conta_str"] = df_pc["CODI_CTA"].astype(str)
            df_pc = df_pc.merge(
                df_saldos[["conta", "saldo"]],
                left_on="conta_str",
                right_on="conta",
                how="left"
            )
            df_pc["Saldo Final"] = df_pc["saldo"].fillna(0.0)
            df_pc = df_pc.drop(columns=["conta_str", "conta", "saldo"], errors="ignore")
        else:
            df_pc["Saldo Final"] = 0.0
        
        self.df_pc = df_pc
        self.df_saldos_finais = df_saldos
        return df_pc
    
    def buscar_movimentacao_periodo(self) -> pd.DataFrame:
        """
        Busca movimentações do período para DRE.
        
        Returns:
            DataFrame com movimentações agrupadas por conta
        """
        df_mov = self.db_client.buscar_movimentacoes_periodo(self.empresa, self.inicio, self.fim)
        self.df_movimentacoes = df_mov
        return df_mov
    
    def buscar_lancamentos_periodo(self) -> pd.DataFrame:
        """
        Busca lançamentos do período para aba de movimentação.
        
        Returns:
            DataFrame com lançamentos do período
        """
        df_lanc = self.db_client.buscar_lancamentos_periodo(self.empresa, self.inicio, self.fim)
        
        # Filtra zeramentos se solicitado
        if self.desconsiderar_zeramento and "orig_lan" in df_lanc.columns:
            df_lanc = df_lanc[df_lanc["orig_lan"] != 2].copy()
        
        # Mapeia contas
        if not df_lanc.empty and self.mapa_codi_to_bc:
            df_lanc["cdeb_lan_str"] = df_lanc["cdeb_lan"].astype(str)
            df_lanc["ccre_lan_str"] = df_lanc["ccre_lan"].astype(str)
            df_lanc["Conta Débito"] = df_lanc["cdeb_lan_str"].map(self.mapa_codi_to_bc)
            df_lanc["Conta Crédito"] = df_lanc["ccre_lan_str"].map(self.mapa_codi_to_bc)
            df_lanc = df_lanc.drop(columns=["cdeb_lan_str", "ccre_lan_str"], errors="ignore")
        
        self.df_lancamentos = df_lanc
        return df_lanc
    
    def gerar_balanco_patrimonial(self) -> pd.DataFrame:
        """
        Gera estrutura do Balanço Patrimonial.
        
        Returns:
            DataFrame com BP estruturado
        """
        if self.df_saldos_finais is None or self.df_saldos_finais.empty:
            return pd.DataFrame()
        
        # Mescla saldos com plano de contas para obter classificação
        df_saldos = self.df_saldos_finais.copy()
        df_pc = self.df_pc.copy()
        
        df_saldos["conta"] = df_saldos["conta"].astype(str)
        df_pc["CODI_CTA_str"] = df_pc["CODI_CTA"].astype(str)
        
        # Mescla
        df_bp = df_saldos.merge(
            df_pc[["CODI_CTA_str", "CLAS_CTA", "NOME_CTA", "BC_GROUP"]],
            left_on="conta",
            right_on="CODI_CTA_str",
            how="left"
        )
        
        # Preenche valores faltantes
        df_bp["NOME_CTA"] = df_bp["NOME_CTA"].fillna("Conta não encontrada")
        df_bp["CLAS_CTA"] = df_bp["CLAS_CTA"].fillna("")
        
        # Classifica contas sem classificação
        df_bp["BC_GROUP"] = df_bp.apply(
            lambda row: self.classificar_beancount(str(row.get("CLAS_CTA", "") or ""), None)
            if pd.isna(row.get("BC_GROUP")) else row["BC_GROUP"],
            axis=1
        )
        
        # Garante que BC_GROUP não seja NaN
        df_bp["BC_GROUP"] = df_bp["BC_GROUP"].fillna("Unknown")
        
        # Agrupa por categoria Beancount
        linhas_bp = []
        
        # Assets (Ativo)
        assets = df_bp[df_bp["BC_GROUP"].str.startswith("Assets", na=False)].copy()
        if not assets.empty:
            linhas_bp.append({"Conta/Categoria": "ATIVO", "Saldo": None})
            
            # Ativo Circulante
            ativo_circ = assets[assets["BC_GROUP"].str.contains("Ativo-Circulante", na=False)]
            if not ativo_circ.empty:
                total_circ = ativo_circ["saldo"].sum()
                linhas_bp.append({"Conta/Categoria": "  Ativo Circulante", "Saldo": total_circ})
                for _, row in ativo_circ.iterrows():
                    linhas_bp.append({
                        "Conta/Categoria": f"    {row['NOME_CTA']} ({row['conta']})",
                        "Saldo": row["saldo"]
                    })
            
            # Ativo Não Circulante
            ativo_ncirc = assets[assets["BC_GROUP"].str.contains("Ativo-Nao-Circulante", na=False)]
            if not ativo_ncirc.empty:
                total_ncirc = ativo_ncirc["saldo"].sum()
                linhas_bp.append({"Conta/Categoria": "  Ativo Não Circulante", "Saldo": total_ncirc})
                for _, row in ativo_ncirc.iterrows():
                    linhas_bp.append({
                        "Conta/Categoria": f"    {row['NOME_CTA']} ({row['conta']})",
                        "Saldo": row["saldo"]
                    })
            
            total_ativo = assets["saldo"].sum()
            linhas_bp.append({"Conta/Categoria": "TOTAL ATIVO", "Saldo": total_ativo})
            linhas_bp.append({"Conta/Categoria": "", "Saldo": None})
        
        # Liabilities (Passivo)
        liabilities = df_bp[df_bp["BC_GROUP"].str.startswith("Liabilities", na=False)].copy()
        if not liabilities.empty:
            linhas_bp.append({"Conta/Categoria": "PASSIVO", "Saldo": None})
            
            # Passivo Circulante
            passivo_circ = liabilities[liabilities["BC_GROUP"].str.contains("Passivo-Circulante", na=False)]
            if not passivo_circ.empty:
                total_circ = passivo_circ["saldo"].sum()
                linhas_bp.append({"Conta/Categoria": "  Passivo Circulante", "Saldo": total_circ})
                for _, row in passivo_circ.iterrows():
                    linhas_bp.append({
                        "Conta/Categoria": f"    {row['NOME_CTA']} ({row['conta']})",
                        "Saldo": row["saldo"]
                    })
            
            # Passivo Não Circulante
            passivo_ncirc = liabilities[liabilities["BC_GROUP"].str.contains("Passivo-Nao-Circulante", na=False)]
            if not passivo_ncirc.empty:
                total_ncirc = passivo_ncirc["saldo"].sum()
                linhas_bp.append({"Conta/Categoria": "  Passivo Não Circulante", "Saldo": total_ncirc})
                for _, row in passivo_ncirc.iterrows():
                    linhas_bp.append({
                        "Conta/Categoria": f"    {row['NOME_CTA']} ({row['conta']})",
                        "Saldo": row["saldo"]
                    })
            
            total_passivo = liabilities["saldo"].sum()
            linhas_bp.append({"Conta/Categoria": "TOTAL PASSIVO", "Saldo": total_passivo})
            linhas_bp.append({"Conta/Categoria": "", "Saldo": None})
        
        # Equity (Patrimônio Líquido)
        equity = df_bp[df_bp["BC_GROUP"].str.startswith("Equity", na=False)].copy()
        if not equity.empty:
            linhas_bp.append({"Conta/Categoria": "PATRIMÔNIO LÍQUIDO", "Saldo": None})
            
            pl_contas = equity[~equity["BC_GROUP"].str.contains("Contas-", na=False)]
            if not pl_contas.empty:
                for _, row in pl_contas.iterrows():
                    linhas_bp.append({
                        "Conta/Categoria": f"  {row['NOME_CTA']} ({row['conta']})",
                        "Saldo": row["saldo"]
                    })
            
            total_pl = equity["saldo"].sum()
            linhas_bp.append({"Conta/Categoria": "TOTAL PATRIMÔNIO LÍQUIDO", "Saldo": total_pl})
            linhas_bp.append({"Conta/Categoria": "", "Saldo": None})
            
            total_geral = (assets["saldo"].sum() if not assets.empty else 0) + \
                         (liabilities["saldo"].sum() if not liabilities.empty else 0) + \
                         total_pl
            linhas_bp.append({"Conta/Categoria": "TOTAL GERAL", "Saldo": total_geral})
        
        return pd.DataFrame(linhas_bp)
    
    def gerar_dre(self) -> pd.DataFrame:
        """
        Gera estrutura da DRE (Demonstração do Resultado do Exercício).
        
        Returns:
            DataFrame com DRE estruturada
        """
        # Garante que o plano de contas está carregado
        if self.df_pc is None:
            self.buscar_plano_contas_com_saldos()
        
        if self.df_movimentacoes is None:
            self.buscar_movimentacao_periodo()
        
        if self.df_movimentacoes is None or self.df_movimentacoes.empty:
            return pd.DataFrame()
        
        # Mescla movimentações com plano de contas para obter classificação
        df_mov = self.df_movimentacoes.copy()
        df_pc = self.df_pc.copy()
        
        # Converte para string para garantir mesclagem correta
        df_mov["conta"] = df_mov["conta"].astype(str).str.strip()
        df_pc["CODI_CTA"] = df_pc["CODI_CTA"].astype(str).str.strip()
        
        # Mescla
        df_dre = df_mov.merge(
            df_pc[["CODI_CTA", "CLAS_CTA", "NOME_CTA", "BC_GROUP"]],
            left_on="conta",
            right_on="CODI_CTA",
            how="left"
        )
        
        # Preenche valores faltantes
        df_dre["NOME_CTA"] = df_dre["NOME_CTA"].fillna("Conta não encontrada")
        df_dre["CLAS_CTA"] = df_dre["CLAS_CTA"].fillna("")
        
        # Classifica contas sem classificação
        df_dre["BC_GROUP"] = df_dre.apply(
            lambda row: self.classificar_beancount(str(row.get("CLAS_CTA", "") or ""), None)
            if pd.isna(row.get("BC_GROUP")) else row["BC_GROUP"],
            axis=1
        )
        
        # Garante que BC_GROUP não seja NaN
        df_dre["BC_GROUP"] = df_dre["BC_GROUP"].fillna("Unknown")
        
        # Debug: alerta sobre contas classificadas como Unknown
        contas_unknown = df_dre[df_dre["BC_GROUP"] == "Unknown"]
        if not contas_unknown.empty:
            print(
                f"\n[DEBUG] {len(contas_unknown)} conta(s) classificada(s) como 'Unknown' na DRE:",
                file=sys.stderr
            )
            for _, row in contas_unknown.iterrows():
                conta = row.get("conta", "?")
                nome = row.get("NOME_CTA", "?")
                clas_cta = row.get("CLAS_CTA", "?")
                movimento = row.get("movimento", 0)
                print(
                    f"  [DEBUG] Conta {conta} | {nome} | CLAS_CTA={clas_cta} | Movimento={movimento:.2f}",
                    file=sys.stderr
                )
        
        linhas_dre = []
        
        # Income (Receitas) - mostra todas as receitas
        income = df_dre[df_dre["BC_GROUP"].str.startswith("Income", na=False)].copy()
        if not income.empty:
            linhas_dre.append({"Item": "RECEITAS", "Valor": None})
            
            # Receitas Operacionais
            rec_op = income[income["BC_GROUP"].str.contains("Operacionais", na=False)]
            if not rec_op.empty:
                total_rec_op = rec_op["movimento"].sum()
                linhas_dre.append({"Item": "  Receitas Operacionais", "Valor": total_rec_op})
                for _, row in rec_op.iterrows():
                    nome_cta = str(row.get("NOME_CTA", "Conta desconhecida"))
                    conta = str(row.get("conta", ""))
                    linhas_dre.append({
                        "Item": f"    {nome_cta} ({conta})",
                        "Valor": row["movimento"]
                    })
                linhas_dre.append({"Item": "  Total Receitas Operacionais", "Valor": total_rec_op})
                linhas_dre.append({"Item": "", "Valor": None})
            
            # Receitas Financeiras
            rec_fin = income[income["BC_GROUP"].str.contains("Financeiras", na=False)]
            if not rec_fin.empty:
                total_rec_fin = rec_fin["movimento"].sum()
                linhas_dre.append({"Item": "  Receitas Financeiras", "Valor": total_rec_fin})
                for _, row in rec_fin.iterrows():
                    nome_cta = str(row.get("NOME_CTA", "Conta desconhecida"))
                    conta = str(row.get("conta", ""))
                    linhas_dre.append({
                        "Item": f"    {nome_cta} ({conta})",
                        "Valor": row["movimento"]
                    })
                linhas_dre.append({"Item": "  Total Receitas Financeiras", "Valor": total_rec_fin})
                linhas_dre.append({"Item": "", "Valor": None})
            
            # Outras Receitas (receitas não classificadas como operacionais ou financeiras)
            outras_rec = income[
                ~income["BC_GROUP"].str.contains("Operacionais", na=False) &
                ~income["BC_GROUP"].str.contains("Financeiras", na=False)
            ]
            if not outras_rec.empty:
                total_outras_rec = outras_rec["movimento"].sum()
                linhas_dre.append({"Item": "  Outras Receitas", "Valor": total_outras_rec})
                for _, row in outras_rec.iterrows():
                    nome_cta = str(row.get("NOME_CTA", "Conta desconhecida"))
                    conta = str(row.get("conta", ""))
                    linhas_dre.append({
                        "Item": f"    {nome_cta} ({conta})",
                        "Valor": row["movimento"]
                    })
                linhas_dre.append({"Item": "  Total Outras Receitas", "Valor": total_outras_rec})
                linhas_dre.append({"Item": "", "Valor": None})
            
            total_receitas = income["movimento"].sum()
            linhas_dre.append({"Item": "TOTAL RECEITAS", "Valor": total_receitas})
            linhas_dre.append({"Item": "", "Valor": None})
        
        # Expenses (Custos e Despesas) - mostra todas as despesas
        expenses = df_dre[df_dre["BC_GROUP"].str.startswith("Expenses", na=False)].copy()
        if not expenses.empty:
            linhas_dre.append({"Item": "(-) CUSTOS E DESPESAS", "Valor": None})
            
            # Custos
            custos = expenses[expenses["BC_GROUP"].str.contains("Custos", na=False)]
            if not custos.empty:
                total_custos = abs(custos["movimento"].sum())
                linhas_dre.append({"Item": "  (-) Custos", "Valor": -total_custos})
                for _, row in custos.iterrows():
                    nome_cta = str(row.get("NOME_CTA", "Conta desconhecida"))
                    conta = str(row.get("conta", ""))
                    linhas_dre.append({
                        "Item": f"    {nome_cta} ({conta})",
                        "Valor": row["movimento"]
                    })
                linhas_dre.append({"Item": "  Total Custos", "Valor": -total_custos})
                linhas_dre.append({"Item": "", "Valor": None})
            
            # Despesas Operacionais
            desp_op = expenses[expenses["BC_GROUP"].str.contains("Despesas-Operacionais", na=False)]
            if not desp_op.empty:
                total_desp_op = abs(desp_op["movimento"].sum())
                linhas_dre.append({"Item": "  (-) Despesas Operacionais", "Valor": -total_desp_op})
                for _, row in desp_op.iterrows():
                    nome_cta = str(row.get("NOME_CTA", "Conta desconhecida"))
                    conta = str(row.get("conta", ""))
                    linhas_dre.append({
                        "Item": f"    {nome_cta} ({conta})",
                        "Valor": row["movimento"]
                    })
                linhas_dre.append({"Item": "  Total Despesas Operacionais", "Valor": -total_desp_op})
                linhas_dre.append({"Item": "", "Valor": None})
            
            # Despesas Financeiras
            desp_fin = expenses[expenses["BC_GROUP"].str.contains("Despesas-Financeiras", na=False)]
            if not desp_fin.empty:
                total_desp_fin = abs(desp_fin["movimento"].sum())
                linhas_dre.append({"Item": "  (-) Despesas Financeiras", "Valor": -total_desp_fin})
                for _, row in desp_fin.iterrows():
                    nome_cta = str(row.get("NOME_CTA", "Conta desconhecida"))
                    conta = str(row.get("conta", ""))
                    linhas_dre.append({
                        "Item": f"    {nome_cta} ({conta})",
                        "Valor": row["movimento"]
                    })
                linhas_dre.append({"Item": "  Total Despesas Financeiras", "Valor": -total_desp_fin})
                linhas_dre.append({"Item": "", "Valor": None})
            
            # Outras Despesas
            outras_desp = expenses[
                ~expenses["BC_GROUP"].str.contains("Custos", na=False) &
                ~expenses["BC_GROUP"].str.contains("Despesas-Operacionais", na=False) &
                ~expenses["BC_GROUP"].str.contains("Despesas-Financeiras", na=False)
            ]
            if not outras_desp.empty:
                total_outras_desp = abs(outras_desp["movimento"].sum())
                linhas_dre.append({"Item": "  (-) Outras Despesas", "Valor": -total_outras_desp})
                for _, row in outras_desp.iterrows():
                    nome_cta = str(row.get("NOME_CTA", "Conta desconhecida"))
                    conta = str(row.get("conta", ""))
                    linhas_dre.append({
                        "Item": f"    {nome_cta} ({conta})",
                        "Valor": row["movimento"]
                    })
                linhas_dre.append({"Item": "  Total Outras Despesas", "Valor": -total_outras_desp})
                linhas_dre.append({"Item": "", "Valor": None})
            
            total_despesas = abs(expenses["movimento"].sum())
            linhas_dre.append({"Item": "TOTAL DESPESAS", "Valor": -total_despesas})
            linhas_dre.append({"Item": "", "Valor": None})
        
        # Resultado
        total_receitas_val = income["movimento"].sum() if not income.empty else 0
        total_despesas_val = abs(expenses["movimento"].sum()) if not expenses.empty else 0
        resultado = total_receitas_val - total_despesas_val
        
        linhas_dre.append({"Item": "RESULTADO DO PERÍODO", "Valor": resultado})
        
        return pd.DataFrame(linhas_dre)
    
    def gerar_balancete(self) -> pd.DataFrame:
        """
        Gera balancete com saldo inicial, débitos, créditos e saldo final.
        
        Returns:
            DataFrame com colunas: Código, Nome, Classificação, 
                                   Saldo Inicial, Total Débitos, Total Créditos, Saldo Final
        """
        # Garante que o plano de contas está carregado
        if self.df_pc is None:
            self.buscar_plano_contas_com_saldos()
        
        # Busca saldos iniciais (até D-1)
        dia_anterior = self.inicio - timedelta(days=1)
        if self.df_saldos_iniciais is None:
            self.df_saldos_iniciais = self.db_client.buscar_saldos(self.empresa, dia_anterior)
        
        # Busca lançamentos do período se ainda não foi carregado
        if self.df_lancamentos is None:
            self.buscar_lancamentos_periodo()
        
        # Inicia com plano de contas
        df_balancete = self.df_pc[["CODI_CTA", "NOME_CTA", "CLAS_CTA"]].copy()
        df_balancete["CODI_CTA"] = df_balancete["CODI_CTA"].astype(str)
        
        # Mescla saldos iniciais
        if self.df_saldos_iniciais is not None and not self.df_saldos_iniciais.empty:
            df_saldos_init = self.df_saldos_iniciais.copy()
            df_saldos_init["conta"] = df_saldos_init["conta"].astype(str)
            df_balancete = df_balancete.merge(
                df_saldos_init[["conta", "saldo"]],
                left_on="CODI_CTA",
                right_on="conta",
                how="left"
            )
            df_balancete["Saldo Inicial"] = df_balancete["saldo"].fillna(0.0)
            df_balancete = df_balancete.drop(columns=["conta", "saldo"], errors="ignore")
        else:
            df_balancete["Saldo Inicial"] = 0.0
        
        # Calcula débitos do período
        if self.df_lancamentos is not None and not self.df_lancamentos.empty:
            # Filtra linhas com débito (cdeb_lan != 0)
            df_debitos = self.df_lancamentos[
                (self.df_lancamentos["cdeb_lan"].astype(str).str.strip() != "0") &
                (self.df_lancamentos["cdeb_lan"].notna())
            ].copy()
            
            if not df_debitos.empty:
                df_debitos["cdeb_lan"] = df_debitos["cdeb_lan"].astype(str).str.strip()
                debitos_agrupados = df_debitos.groupby("cdeb_lan")["vlor_lan"].sum().reset_index()
                debitos_agrupados.columns = ["conta", "Total Débitos"]
                debitos_agrupados["conta"] = debitos_agrupados["conta"].astype(str).str.strip()
                df_balancete = df_balancete.merge(
                    debitos_agrupados,
                    left_on="CODI_CTA",
                    right_on="conta",
                    how="left"
                )
                df_balancete["Total Débitos"] = df_balancete["Total Débitos"].fillna(0.0)
                df_balancete = df_balancete.drop(columns=["conta"], errors="ignore")
            else:
                df_balancete["Total Débitos"] = 0.0
            
            # Calcula créditos do período
            # Filtra linhas com crédito (ccre_lan != 0)
            df_creditos = self.df_lancamentos[
                (self.df_lancamentos["ccre_lan"].astype(str).str.strip() != "0") &
                (self.df_lancamentos["ccre_lan"].notna())
            ].copy()
            
            if not df_creditos.empty:
                df_creditos["ccre_lan"] = df_creditos["ccre_lan"].astype(str).str.strip()
                creditos_agrupados = df_creditos.groupby("ccre_lan")["vlor_lan"].sum().reset_index()
                creditos_agrupados.columns = ["conta", "Total Créditos"]
                creditos_agrupados["conta"] = creditos_agrupados["conta"].astype(str).str.strip()
                df_balancete = df_balancete.merge(
                    creditos_agrupados,
                    left_on="CODI_CTA",
                    right_on="conta",
                    how="left"
                )
                df_balancete["Total Créditos"] = df_balancete["Total Créditos"].fillna(0.0)
                df_balancete = df_balancete.drop(columns=["conta"], errors="ignore")
            else:
                df_balancete["Total Créditos"] = 0.0
        else:
            df_balancete["Total Débitos"] = 0.0
            df_balancete["Total Créditos"] = 0.0
        
        # Calcula saldo final = saldo inicial + débitos - créditos
        df_balancete["Saldo Final"] = (
            df_balancete["Saldo Inicial"] + 
            df_balancete["Total Débitos"] - 
            df_balancete["Total Créditos"]
        )
        
        # Arredonda valores para 2 casas decimais
        df_balancete["Saldo Inicial"] = df_balancete["Saldo Inicial"].round(2)
        df_balancete["Total Débitos"] = df_balancete["Total Débitos"].round(2)
        df_balancete["Total Créditos"] = df_balancete["Total Créditos"].round(2)
        df_balancete["Saldo Final"] = df_balancete["Saldo Final"].round(2)
        
        # Renomeia colunas para o formato final
        df_balancete = df_balancete.rename(columns={
            "CODI_CTA": "Código",
            "NOME_CTA": "Nome",
            "CLAS_CTA": "Classificação"
        })
        
        # Ordena por classificação
        df_balancete = df_balancete.sort_values("Classificação")
        
        # Seleciona apenas colunas necessárias na ordem correta
        df_balancete = df_balancete[[
            "Código", "Nome", "Classificação", 
            "Saldo Inicial", "Total Débitos", "Total Créditos", "Saldo Final"
        ]]
        
        return df_balancete
    
    def _aplicar_formatacao(self, ws, num_cols: int, num_rows: int, coluna_codigo_texto: Optional[int] = None, colunas_texto: Optional[List[int]] = None):
        """
        Aplica formatação básica à planilha.
        
        Args:
            ws: Worksheet do openpyxl
            num_cols: Número de colunas
            num_rows: Número de linhas
            coluna_codigo_texto: Coluna que deve ser formatada como texto (sem formatação numérica) - deprecated, use colunas_texto
            colunas_texto: Lista de colunas (1-indexed) que devem ser formatadas como texto
        """
        # Compatibilidade: se coluna_codigo_texto for fornecido, adiciona à lista
        if colunas_texto is None:
            colunas_texto = []
        if coluna_codigo_texto is not None and coluna_codigo_texto not in colunas_texto:
            colunas_texto.append(coluna_codigo_texto)
        # Estilos
        font_titulo = Font(bold=True, size=12)
        font_subtitulo = Font(bold=True, size=11)
        font_normal = Font(size=10)
        align_right = Alignment(horizontal="right")
        align_left = Alignment(horizontal="left")
        
        # Formato de moeda brasileiro
        moeda_format = '#,##0.00'  # Será aplicado via número
        
        # Borda
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplica formatação
        for row in range(1, min(num_rows + 1, ws.max_row + 1)):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                
                # Formatação de texto
                if row == 1:  # Cabeçalho
                    cell.font = font_titulo
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.value and isinstance(cell.value, str):
                    # Títulos e subtotais em negrito
                    if any(keyword in str(cell.value).upper() for keyword in ["TOTAL", "ATIVO", "PASSIVO", "PATRIMÔNIO", "RECEITAS", "DESPESAS", "CUSTOS", "RESULTADO"]):
                        cell.font = font_subtitulo
                    else:
                        cell.font = font_normal
                    cell.alignment = align_left
                
                # Formatação numérica (exceto para colunas de texto especificadas)
                if col in colunas_texto and cell.value is not None:
                    # Colunas de texto: mantém como texto, sem formatação numérica
                    cell.number_format = '@'  # Formato texto
                    cell.alignment = align_left
                elif col not in colunas_texto and isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                
                # Borda
                cell.border = thin_border
        
        # Autoajusta largura das colunas
        for col in range(1, num_cols + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def exportar_excel(self, outdir: Path, nome_arquivo: Optional[str] = None) -> Path:
        """
        Exporta dados contábeis para arquivo Excel.
        
        Args:
            outdir: Diretório de saída
            nome_arquivo: Nome do arquivo (opcional, será gerado automaticamente se None)
            
        Returns:
            Caminho do arquivo Excel gerado
        """
        # Cria diretório se não existir
        outdir = Path(outdir)
        outdir.mkdir(parents=True, exist_ok=True)
        
        # Nome do arquivo
        if nome_arquivo is None:
            nome_arquivo = f"contabilidade_{self.empresa}_{self.inicio}_{self.fim}.xlsx"
        
        excel_path = outdir / nome_arquivo
        
        # Busca dados
        if self.df_pc is None:
            self.buscar_plano_contas_com_saldos()
        
        if self.df_movimentacoes is None:
            self.buscar_movimentacao_periodo()
        
        if self.df_lancamentos is None:
            self.buscar_lancamentos_periodo()
        
        # Cria workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove planilha padrão
        
        # Aba 1: Plano de Contas
        if self.df_pc is not None and not self.df_pc.empty:
            ws_pc = wb.create_sheet("Plano de Contas")
            df_pc_export = self.df_pc[["CODI_CTA", "NOME_CTA", "CLAS_CTA", "TIPO_CTA", "SITUACAO_CTA", "BC_ACCOUNT"]].copy()
            df_pc_export = df_pc_export.sort_values("CLAS_CTA")
            
            # Cabeçalho
            headers = ["Código", "Nome", "Classificação", "Tipo", "Situação", "Classificação Beancount"]
            ws_pc.append(headers)
            
            # Dados
            for _, row in df_pc_export.iterrows():
                # Código como texto para evitar formatação numérica
                codigo = str(row["CODI_CTA"]) if pd.notna(row["CODI_CTA"]) else ""
                ws_pc.append([
                    codigo,
                    row["NOME_CTA"],
                    row["CLAS_CTA"],
                    row["TIPO_CTA"],
                    row["SITUACAO_CTA"],
                    row["BC_ACCOUNT"]
                ])
            
            self._aplicar_formatacao(ws_pc, len(headers), len(df_pc_export) + 1, coluna_codigo_texto=1)
        
        # Aba 2: Balanço Patrimonial
        df_bp = self.gerar_balanco_patrimonial()
        if not df_bp.empty:
            ws_bp = wb.create_sheet("Balanço Patrimonial")
            headers = ["Conta/Categoria", "Saldo"]
            ws_bp.append(headers)
            
            for _, row in df_bp.iterrows():
                ws_bp.append([row["Conta/Categoria"], row["Saldo"]])
            
            self._aplicar_formatacao(ws_bp, len(headers), len(df_bp) + 1)
        
        # Aba 3: DRE
        df_dre = self.gerar_dre()
        if not df_dre.empty:
            ws_dre = wb.create_sheet("DRE")
            headers = ["Item", "Valor"]
            ws_dre.append(headers)
            
            for _, row in df_dre.iterrows():
                ws_dre.append([row["Item"], row["Valor"]])
            
            self._aplicar_formatacao(ws_dre, len(headers), len(df_dre) + 1)
        
        # Aba 4: Movimentação do Período
        if self.df_lancamentos is not None and not self.df_lancamentos.empty:
            ws_mov = wb.create_sheet("Movimentação do Período")
            df_mov_export = self.df_lancamentos.copy()
            
            # Garante que as colunas necessárias existam
            if "Conta Débito" not in df_mov_export.columns:
                df_mov_export["Conta Débito"] = ""
            if "Conta Crédito" not in df_mov_export.columns:
                df_mov_export["Conta Crédito"] = ""
            
            # Adiciona colunas de código (cdeb_lan e ccre_lan devem estar no DataFrame)
            if "cdeb_lan" not in df_mov_export.columns:
                df_mov_export["cdeb_lan"] = ""
            if "ccre_lan" not in df_mov_export.columns:
                df_mov_export["ccre_lan"] = ""
            
            # Cria colunas de código formatadas como string
            def limpar_codigo(valor):
                if pd.isna(valor):
                    return ""
                str_val = str(valor).strip()
                if str_val in ["0", "nan", "None", ""]:
                    return ""
                return str_val
            
            df_mov_export["Código Débito"] = df_mov_export["cdeb_lan"].apply(limpar_codigo)
            df_mov_export["Código Crédito"] = df_mov_export["ccre_lan"].apply(limpar_codigo)
            
            # Seleciona colunas disponíveis
            colunas_disponiveis = []
            for col in ["data_lan", "Código Débito", "Conta Débito", "Código Crédito", "Conta Crédito", "chis_lan", "ndoc_lan", "codi_lote", "vlor_lan"]:
                if col in df_mov_export.columns:
                    colunas_disponiveis.append(col)
            
            df_mov_export = df_mov_export[colunas_disponiveis].copy()
            
            # Ordena se possível
            if "data_lan" in df_mov_export.columns:
                if "codi_lote" in df_mov_export.columns:
                    df_mov_export = df_mov_export.sort_values(["data_lan", "codi_lote"])
                else:
                    df_mov_export = df_mov_export.sort_values("data_lan")
            
            # Cabeçalhos: Data, Código Débito, Conta Débito, Código Crédito, Conta Crédito, Histórico, Documento, Lote, Valor
            headers = ["Data", "Código Débito", "Conta Débito", "Código Crédito", "Conta Crédito", "Histórico", "Documento", "Lote", "Valor"]
            ws_mov.append(headers)
            
            # Função helper para converter números inteiros corretamente
            def formatar_numero_texto(valor):
                """Converte valor numérico para string, removendo .0 de inteiros."""
                if pd.isna(valor):
                    return ""
                # Se for numérico e inteiro, converte para int primeiro
                if isinstance(valor, (int, float)):
                    if float(valor).is_integer():
                        return str(int(valor))
                    return str(valor)
                str_val = str(valor).strip()
                if str_val in ["nan", "None", ""]:
                    return ""
                return str_val
            
            for _, row in df_mov_export.iterrows():
                # Converte codi_lote e ndoc_lan para string (formato texto)
                codi_lote_val = row.get("codi_lote", "")
                codi_lote_str = formatar_numero_texto(codi_lote_val)
                if codi_lote_str == "0":
                    codi_lote_str = ""
                
                ndoc_lan_val = row.get("ndoc_lan", "")
                ndoc_lan_str = formatar_numero_texto(ndoc_lan_val)
                
                ws_mov.append([
                    row.get("data_lan", ""),
                    row.get("Código Débito", ""),
                    row.get("Conta Débito", ""),
                    row.get("Código Crédito", ""),
                    row.get("Conta Crédito", ""),
                    row.get("chis_lan", ""),
                    ndoc_lan_str,
                    codi_lote_str,
                    row.get("vlor_lan", 0)
                ])
            
            # Colunas de texto: 2 (Código Débito), 4 (Código Crédito), 7 (Documento), 8 (Lote)
            colunas_texto = [2, 4, 7, 8]
            self._aplicar_formatacao(ws_mov, len(headers), len(df_mov_export) + 1, colunas_texto=colunas_texto)
        
        # Aba 5: Balancete
        df_balancete = self.gerar_balancete()
        if not df_balancete.empty:
            ws_balancete = wb.create_sheet("Balancete")
            headers = ["Código", "Nome", "Classificação", "Saldo Inicial", "Total Débitos", "Total Créditos", "Saldo Final"]
            ws_balancete.append(headers)
            
            for _, row in df_balancete.iterrows():
                # Código como texto para evitar formatação numérica
                codigo = str(row["Código"]) if pd.notna(row["Código"]) else ""
                ws_balancete.append([
                    codigo,
                    row["Nome"],
                    row["Classificação"],
                    row["Saldo Inicial"],
                    row["Total Débitos"],
                    row["Total Créditos"],
                    row["Saldo Final"]
                ])
            
            self._aplicar_formatacao(ws_balancete, len(headers), len(df_balancete) + 1, coluna_codigo_texto=1)
        
        # Salva arquivo
        wb.save(excel_path)
        
        return excel_path

