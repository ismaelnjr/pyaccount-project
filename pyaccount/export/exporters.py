#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo para exportação de dados contábeis.

Classes base para exportação:
- BeancountExporter: Exporta dados para formato Beancount
- ExcelExporter: Exporta dados para formato Excel
"""
from pathlib import Path
from datetime import date, timedelta
from typing import Dict, Optional, List
import sys
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pyaccount.data.db_client import ContabilDBClient
from pyaccount.core.account_classifier import AccountClassifier, TipoPlanoContas, obter_classificacao_do_modelo
from pyaccount.core.account_mapper import AccountMapper
from pyaccount.builders.financial_statements import (
    BalanceSheetBuilder,
    IncomeStatementBuilder,
    TrialBalanceBuilder,
    PeriodMovementsBuilder
)
from pyaccount.core.utils import fmt_amount, normalizar_nome


class BeancountExporter:
    """
    Exportador de dados contábeis para formato Beancount.
    
    Gera arquivo .beancount formatado com:
    - Declarações open das contas
    - Transação de abertura (saldos até D-1)
    - Lançamentos do período agrupados por lote
    """
    
    def __init__(
        self,
        df_saldos: pd.DataFrame,
        df_lancamentos: pd.DataFrame,
        mapa_codi_to_bc: Dict[str, str],
        empresa: int,
        inicio: date,
        fim: date,
        moeda: str,
        abrir_equity_abertura: str = "Equity:Abertura"
    ):
        """
        Inicializa o exportador Beancount.
        
        Args:
            df_saldos: DataFrame com saldos de abertura (deve ter colunas: BC_ACCOUNT, saldo)
            df_lancamentos: DataFrame com lançamentos do período (deve ter colunas: codi_lote, data_lan, 
                           cdeb_lan, ccre_lan, vlor_lan, chis_lan, ndoc_lan, codi_usu, BC_DEB, BC_CRE)
            mapa_codi_to_bc: Dicionário mapeando CODI_CTA -> BC_ACCOUNT
            empresa: Código da empresa
            inicio: Data inicial do período
            fim: Data final do período
            moeda: Código da moeda (ex: "BRL")
            abrir_equity_abertura: Nome da conta Equity para transação de abertura
        """
        self.df_saldos = df_saldos
        self.df_lancamentos = df_lancamentos
        self.mapa_codi_to_bc = mapa_codi_to_bc
        self.empresa = empresa
        self.inicio = inicio
        self.fim = fim
        self.moeda = moeda
        self.abrir_equity_abertura = abrir_equity_abertura
    
    def exportar(self, caminho: Path) -> Path:
        """
        Exporta dados para arquivo Beancount.
        
        Args:
            caminho: Caminho do arquivo de saída
            
        Returns:
            Caminho do arquivo gerado
        """
        caminho.parent.mkdir(parents=True, exist_ok=True)
        dia_anterior = self.inicio - timedelta(days=1)
        
        # Coleta todas as contas usadas
        contas_usadas = set()
        if self.df_saldos is not None and not self.df_saldos.empty:
            contas_usadas.update(self.df_saldos["BC_ACCOUNT"].tolist())
        if self.df_lancamentos is not None and not self.df_lancamentos.empty:
            contas_usadas.update(self.df_lancamentos["BC_DEB"].dropna().tolist())
            contas_usadas.update(self.df_lancamentos["BC_CRE"].dropna().tolist())
        contas_usadas.add(self.abrir_equity_abertura)
        
        # Escreve arquivo Beancount
        with caminho.open("w", encoding="utf-8") as f:
            # Cabeçalho
            self._escrever_cabecalho(f)
            
            # Declarações open
            self._escrever_opens(f, contas_usadas)
            
            # Transação de abertura
            self._escrever_transacao_abertura(f, dia_anterior)
            
            # Lançamentos agrupados por lote
            self._escrever_lancamentos(f)
        
        return caminho
    
    def _escrever_cabecalho(self, f) -> None:
        """Escreve cabeçalho do arquivo Beancount."""
        f.write(f"; Empresa {self.empresa} — período {self.inicio} a {self.fim}\n")
        f.write(f'option "operating_currency" "{self.moeda}"\n')
        f.write('option "title" "Contabilidade — Extração ODBC"\n\n')
    
    def _escrever_opens(self, f, contas_usadas: set) -> None:
        """Escreve declarações open das contas."""
        for acc in sorted(contas_usadas):
            f.write(f"{self.inicio} open {acc} {self.moeda}\n")
        f.write("\n")
    
    def _escrever_transacao_abertura(self, f, dia_anterior: date) -> None:
        """Escreve transação de abertura."""
        if self.df_saldos is not None and not self.df_saldos.empty:
            f.write(f'{self.inicio} * "Abertura de saldos" "Saldo até {dia_anterior}"\n')
            for _, r in self.df_saldos.iterrows():
                f.write(f"  {r['BC_ACCOUNT']:<60} {fmt_amount(r['saldo'], self.moeda)}\n")
            f.write(f"  {self.abrir_equity_abertura}\n\n")
    
    def _escrever_lancamentos(self, f) -> None:
        """Escreve lançamentos agrupados por lote."""
        if self.df_lancamentos is None or self.df_lancamentos.empty:
            return
        
        # Agrupa por codi_lote e data_lan
        df_lanc_filtrado = self.df_lancamentos[
            (self.df_lancamentos["cdeb_lan"].astype(str).str.strip() != "0") |
            (self.df_lancamentos["ccre_lan"].astype(str).str.strip() != "0")
        ].copy()
        
        for (lote_id, data_lan), grupo in df_lanc_filtrado.groupby(["codi_lote", "data_lan"]):
            # Processa débitos: filtra linhas onde cdeb_lan != 0 e BC_DEB não é NaN
            debitos_df = grupo[
                (grupo["cdeb_lan"].astype(str).str.strip() != "0") &
                (grupo["BC_DEB"].notna())
            ].copy()
            
            # Processa créditos: filtra linhas onde ccre_lan != 0 e BC_CRE não é NaN
            creditos_df = grupo[
                (grupo["ccre_lan"].astype(str).str.strip() != "0") &
                (grupo["BC_CRE"].notna())
            ].copy()
            
            # Agrupa débitos por conta e soma valores
            debitos_por_conta = {}
            if not debitos_df.empty:
                for conta_deb, subgrupo in debitos_df.groupby("BC_DEB"):
                    debitos_por_conta[conta_deb] = float(subgrupo["vlor_lan"].sum())
            
            # Agrupa créditos por conta e soma valores
            creditos_por_conta = {}
            if not creditos_df.empty:
                for conta_cre, subgrupo in creditos_df.groupby("BC_CRE"):
                    creditos_por_conta[conta_cre] = float(subgrupo["vlor_lan"].sum())
            
            # Valida que soma de débitos = soma de créditos
            total_debitos = sum(debitos_por_conta.values())
            total_creditos = sum(creditos_por_conta.values())
            
            # Ignora lotes sem débitos ou créditos válidos, ou que não estão balanceados
            if not debitos_por_conta and not creditos_por_conta:
                continue
            
            if abs(total_debitos - total_creditos) > 0.01:
                # Detecta contas de débito não mapeadas
                contas_debito_sem_map = grupo[
                    (grupo["cdeb_lan"].astype(str).str.strip() != "0") &
                    (grupo["BC_DEB"].isna())
                ]
                debitos_nao_encontrados = []
                if not contas_debito_sem_map.empty:
                    debitos_nao_encontrados = contas_debito_sem_map["cdeb_lan"].unique().tolist()
                    debitos_nao_encontrados = [str(int(c)) if pd.notna(c) else "?" for c in debitos_nao_encontrados]
                
                # Detecta contas de crédito não mapeadas
                contas_credito_sem_map = grupo[
                    (grupo["ccre_lan"].astype(str).str.strip() != "0") &
                    (grupo["BC_CRE"].isna())
                ]
                creditos_nao_encontrados = []
                if not contas_credito_sem_map.empty:
                    creditos_nao_encontrados = contas_credito_sem_map["ccre_lan"].unique().tolist()
                    creditos_nao_encontrados = [str(int(c)) if pd.notna(c) else "?" for c in creditos_nao_encontrados]
                
                # Monta mensagem de aviso com detalhes
                msg = (
                    f"[aviso] Lote {lote_id} não balanceado: "
                    f"débitos={total_debitos:.2f}, créditos={total_creditos:.2f}"
                )
                
                detalhes = []
                if debitos_nao_encontrados:
                    detalhes.append(f"Débito(s) não encontrado(s): {', '.join(debitos_nao_encontrados)}")
                if creditos_nao_encontrados:
                    detalhes.append(f"Crédito(s) não encontrado(s): {', '.join(creditos_nao_encontrados)}")
                
                if detalhes:
                    msg += " | " + " | ".join(detalhes)
                
                print(msg, file=sys.stderr)
                continue
            
            # Obtém metadados do primeiro registro do lote
            primeiro_registro = grupo.iloc[0]
            data_txt = data_lan.strftime("%Y-%m-%d")
            hist = (str(primeiro_registro.get('chis_lan') or '')).replace('\\n', ' ').strip()
            ndoc = str(primeiro_registro.get('ndoc_lan') or '')
            lote = str(lote_id)
            usu = str(primeiro_registro.get('codi_usu') or '')
            meta = " ".join(filter(None, [
                f'Doc {ndoc}' if ndoc and ndoc != 'nan' else '', 
                f'Lote {lote}' if lote and lote != 'nan' else '', 
                f'Usu {usu}' if usu and usu != 'nan' else ''
            ]))
            
            # Escreve cabeçalho da transação
            f.write(f'{data_txt} * "{hist}" "{meta}"\n')
            
            # Escreve linhas de débito (positivas)
            for conta_deb, valor in sorted(debitos_por_conta.items()):
                f.write(f"  {conta_deb:<60} {fmt_amount(valor, self.moeda)}\n")
            
            # Escreve linhas de crédito (negativas)
            for conta_cre, valor in sorted(creditos_por_conta.items()):
                f.write(f"  {conta_cre:<60} {fmt_amount(-valor, self.moeda)}\n")
            
            f.write("\n")


class ExcelExporter:
    """
    Exportador de dados contábeis para Excel.
    
    Esta classe encapsula a lógica necessária para:
    - Buscar plano de contas com saldos finais
    - Gerar Balanço Patrimonial estruturado
    - Gerar DRE com cálculos
    - Exportar movimentações do período
    - Formatar e exportar tudo para Excel
    """
    
    def __init__(
        self,
        db_client: ContabilDBClient,
        empresa: int,
        inicio: date,
        fim: date,
        classificacao_customizada: Optional[Dict[str, str]] = None,
        modelo: Optional[TipoPlanoContas] = None,
        desconsiderar_zeramento: bool = True,
        agrupamento_periodo: Optional[str] = None,
    ):
        """
        Inicializa o exportador Excel.
        
        Args:
            db_client: Cliente de banco de dados
            empresa: Código da empresa
            inicio: Data inicial do período
            fim: Data final do período
            classificacao_customizada: Dicionário opcional com mapeamento customizado
                                      de prefixos CLAS_CTA para categorias Beancount.
                                      Se fornecido, tem prioridade sobre o modelo.
            modelo: Tipo de plano de contas a usar. Se None, usa CLASSIFICACAO_PADRAO.
            desconsiderar_zeramento: Se True, exclui lançamentos com orig_lan = 2 (Zeramento)
            agrupamento_periodo: Tipo de agrupamento por período na DRE. Valores aceitos:
                                None (sem agrupamento), "mensal" ou "trimestral"
        """
        # Valida agrupamento_periodo
        if agrupamento_periodo is not None and agrupamento_periodo not in ["mensal", "trimestral"]:
            raise ValueError(
                f"agrupamento_periodo deve ser None, 'mensal' ou 'trimestral'. "
                f"Recebido: {agrupamento_periodo}"
            )
        
        self.db_client = db_client
        self.empresa = empresa
        self.inicio = inicio
        self.fim = fim
        self.classificacao_customizada = classificacao_customizada
        self.modelo = modelo
        self.desconsiderar_zeramento = desconsiderar_zeramento
        self.agrupamento_periodo = agrupamento_periodo
        
        # Obtém classificação baseada no modelo e customizações
        classificacao = obter_classificacao_do_modelo(modelo, classificacao_customizada)
        
        # Mapeador de contas (classe base compartilhada)
        self.account_mapper = AccountMapper(classificacao)
        
        # DataFrames internos
        self.df_pc: Optional[pd.DataFrame] = None
        self.df_saldos_finais: Optional[pd.DataFrame] = None
        self.df_saldos_iniciais: Optional[pd.DataFrame] = None
        self.df_movimentacoes: Optional[pd.DataFrame] = None
        self.df_lancamentos: Optional[pd.DataFrame] = None
        self.mapa_codi_to_bc: Dict[str, str] = {}
    
    def _criar_mapa_tipo_conta(self) -> Dict[str, str]:
        """
        Cria mapa de código da conta -> TIPO_CTA a partir do plano de contas.
        
        Returns:
            Dicionário mapeando código da conta (str) -> TIPO_CTA ("S" ou "A")
        """
        mapa = {}
        if self.df_pc is not None and not self.df_pc.empty:
            if "CODI_CTA" in self.df_pc.columns and "TIPO_CTA" in self.df_pc.columns:
                for _, row in self.df_pc.iterrows():
                    codi_cta = str(row["CODI_CTA"]).strip()
                    tipo_cta = str(row["TIPO_CTA"]).strip() if pd.notna(row["TIPO_CTA"]) else ""
                    if codi_cta and tipo_cta:
                        mapa[codi_cta] = tipo_cta
        return mapa
    
    def classificar_beancount(self, clas_cta: str, tipo_cta: Optional[str] = None) -> str:
        """Mapeia CLAS_CTA -> grupo Beancount."""
        return self.account_mapper.classificar_beancount(clas_cta, tipo_cta)
    
    def buscar_plano_contas_com_saldos(self) -> pd.DataFrame:
        """
        Busca plano de contas do banco de dados com saldos finais.
        
        Returns:
            DataFrame com plano de contas e saldos finais
        """
        # Busca plano de contas
        df_pc = self.db_client.buscar_plano_contas(self.empresa)
        
        if df_pc.empty:
            raise RuntimeError("Plano de contas vazio para a empresa informada.")
        
        # Processa plano de contas usando AccountMapper
        df_pc = self.account_mapper.processar_plano_contas(df_pc, filtrar_ativas=False)
        
        # Cria mapa para lookup
        mapas = self.account_mapper.criar_mapas(df_pc)
        self.mapa_codi_to_bc = mapas["codi_to_bc"]
        
        # Busca saldos finais
        df_saldos = self.db_client.buscar_saldos(self.empresa, self.fim)
        
        # Mescla saldos ao plano de contas
        if not df_saldos.empty:
            df_saldos["conta"] = df_saldos["conta"].astype(str)
            df_pc["conta_str"] = df_pc["CODI_CTA"].astype(str)
            df_pc = df_pc.merge(
                df_saldos[["conta", "saldo"]],
                left_on="conta_str",
                right_on="conta",
                how="left"
            )
            df_pc["Saldo Final"] = df_pc["saldo"].fillna(0.0)
            df_pc = df_pc.drop(columns=["conta_str", "conta", "saldo"], errors="ignore")
        else:
            df_pc["Saldo Final"] = 0.0
        
        self.df_pc = df_pc
        self.df_saldos_finais = df_saldos
        return df_pc
    
    def buscar_movimentacao_periodo(self) -> pd.DataFrame:
        """
        Busca movimentações do período para DRE.
        
        Returns:
            DataFrame com movimentações agrupadas por conta
        """
        df_mov = self.db_client.buscar_movimentacoes_periodo(self.empresa, self.inicio, self.fim)
        self.df_movimentacoes = df_mov
        return df_mov
    
    def _calcular_movimentacoes_por_periodo(self) -> pd.DataFrame:
        """
        Calcula movimentações por conta e período (mensal ou trimestral).
        
        Returns:
            DataFrame com colunas: conta, periodo, movimento
        """
        # Busca lançamentos detalhados do período
        df_lanc = self.db_client.buscar_lancamentos_periodo(self.empresa, self.inicio, self.fim)
        
        # Filtra zeramentos se solicitado
        if self.desconsiderar_zeramento and "orig_lan" in df_lanc.columns:
            df_lanc = df_lanc[df_lanc["orig_lan"] != 2].copy()
        
        if df_lanc.empty:
            return pd.DataFrame(columns=["conta", "periodo", "movimento"])
        
        # Converte data_lan para datetime se necessário
        if not pd.api.types.is_datetime64_any_dtype(df_lanc["data_lan"]):
            df_lanc["data_lan"] = pd.to_datetime(df_lanc["data_lan"])
        
        # Calcula período baseado no tipo de agrupamento
        if self.agrupamento_periodo == "mensal":
            # Formato: "Jan/24", "Fev/24", etc.
            df_lanc["periodo"] = df_lanc["data_lan"].dt.strftime("%b/%y").str.title()
        elif self.agrupamento_periodo == "trimestral":
            # Formato: "1T/24", "2T/24", etc.
            df_lanc["trimestre"] = df_lanc["data_lan"].dt.quarter
            df_lanc["ano"] = df_lanc["data_lan"].dt.strftime("%y")
            df_lanc["periodo"] = df_lanc["trimestre"].astype(str) + "T/" + df_lanc["ano"]
            df_lanc = df_lanc.drop(columns=["trimestre", "ano"], errors="ignore")
        else:
            # Não deveria chegar aqui, mas para segurança
            df_lanc["periodo"] = "Total"
        
        # Prepara dados para cálculo de movimentações
        # Débitos: valores positivos (aumentam saldo)
        df_debitos = df_lanc[
            (df_lanc["cdeb_lan"].astype(str).str.strip() != "0") &
            (df_lanc["cdeb_lan"].notna())
        ].copy()
        
        # Créditos: valores negativos (diminuem saldo)
        df_creditos = df_lanc[
            (df_lanc["ccre_lan"].astype(str).str.strip() != "0") &
            (df_lanc["ccre_lan"].notna())
        ].copy()
        
        # Converte contas para string
        if not df_debitos.empty:
            df_debitos["conta"] = df_debitos["cdeb_lan"].astype(str).str.strip()
            df_debitos["movimento"] = df_debitos["vlor_lan"]
        else:
            df_debitos = pd.DataFrame(columns=["conta", "periodo", "movimento"])
        
        if not df_creditos.empty:
            df_creditos["conta"] = df_creditos["ccre_lan"].astype(str).str.strip()
            df_creditos["movimento"] = -df_creditos["vlor_lan"]  # Negativo para créditos
        else:
            df_creditos = pd.DataFrame(columns=["conta", "periodo", "movimento"])
        
        # Combina débitos e créditos
        df_movimentos = pd.concat([
            df_debitos[["conta", "periodo", "movimento"]],
            df_creditos[["conta", "periodo", "movimento"]]
        ], ignore_index=True)
        
        # Agrupa por conta e período
        df_result = df_movimentos.groupby(["conta", "periodo"], as_index=False)["movimento"].sum()
        
        # Remove períodos com movimento zero
        df_result = df_result[df_result["movimento"] != 0].copy()
        
        return df_result
    
    def buscar_lancamentos_periodo(self) -> pd.DataFrame:
        """
        Busca lançamentos do período para aba de movimentação.
        
        Returns:
            DataFrame com lançamentos do período
        """
        df_lanc = self.db_client.buscar_lancamentos_periodo(self.empresa, self.inicio, self.fim)
        
        # Filtra zeramentos se solicitado
        if self.desconsiderar_zeramento and "orig_lan" in df_lanc.columns:
            df_lanc = df_lanc[df_lanc["orig_lan"] != 2].copy()
        
        # Mapeia contas
        if not df_lanc.empty and self.mapa_codi_to_bc:
            df_lanc["cdeb_lan_str"] = df_lanc["cdeb_lan"].astype(str)
            df_lanc["ccre_lan_str"] = df_lanc["ccre_lan"].astype(str)
            df_lanc["Conta Débito"] = df_lanc["cdeb_lan_str"].map(self.mapa_codi_to_bc)
            df_lanc["Conta Crédito"] = df_lanc["ccre_lan_str"].map(self.mapa_codi_to_bc)
            df_lanc = df_lanc.drop(columns=["cdeb_lan_str", "ccre_lan_str"], errors="ignore")
        
        self.df_lancamentos = df_lanc
        return df_lanc
    
    def gerar_balanco_patrimonial(self) -> pd.DataFrame:
        """
        Gera estrutura do Balanço Patrimonial.
        
        Returns:
            DataFrame com BP estruturado
        """
        if self.df_saldos_finais is None or self.df_saldos_finais.empty:
            return pd.DataFrame()
        
        builder = BalanceSheetBuilder(
            self.df_saldos_finais,
            self.df_pc,
            self.account_mapper
        )
        return builder.gerar()
    
    def gerar_dre(self) -> pd.DataFrame:
        """
        Gera estrutura da DRE (Demonstração do Resultado do Exercício).
        
        Returns:
            DataFrame com DRE estruturada. Se agrupamento_periodo estiver definido,
            retorna DataFrame com colunas: Item + colunas dinâmicas por período + Total.
            Caso contrário, retorna DataFrame com colunas: Item, Valor.
        """
        # Garante que o plano de contas está carregado
        if self.df_pc is None:
            self.buscar_plano_contas_com_saldos()
        
        # Se há agrupamento por período, usa método específico
        if self.agrupamento_periodo:
            df_mov_por_periodo = self._calcular_movimentacoes_por_periodo()
            if df_mov_por_periodo.empty:
                return pd.DataFrame()
            
            builder = IncomeStatementBuilder(
                df_mov_por_periodo,
                self.df_pc,
                self.account_mapper,
                agrupamento_periodo=self.agrupamento_periodo
            )
            return builder.gerar()
        
        # Comportamento padrão (sem agrupamento)
        if self.df_movimentacoes is None:
            self.buscar_movimentacao_periodo()
        
        if self.df_movimentacoes is None or self.df_movimentacoes.empty:
            return pd.DataFrame()
        
        builder = IncomeStatementBuilder(
            self.df_movimentacoes,
            self.df_pc,
            self.account_mapper
        )
        return builder.gerar()
    
    def gerar_balancete(self) -> pd.DataFrame:
        """
        Gera balancete com saldo inicial, débitos, créditos e saldo final.
        
        Returns:
            DataFrame com colunas: Código, Nome, Classificação, 
                                   Saldo Inicial, Total Débitos, Total Créditos, Saldo Final
        """
        # Garante que o plano de contas está carregado
        if self.df_pc is None:
            self.buscar_plano_contas_com_saldos()
        
        # Busca saldos iniciais (até D-1)
        dia_anterior = self.inicio - timedelta(days=1)
        if self.df_saldos_iniciais is None:
            self.df_saldos_iniciais = self.db_client.buscar_saldos(self.empresa, dia_anterior)
        
        # Busca lançamentos do período se ainda não foi carregado
        if self.df_lancamentos is None:
            self.buscar_lancamentos_periodo()
        
        builder = TrialBalanceBuilder(
            self.df_pc,
            self.df_saldos_iniciais,
            self.df_lancamentos,
            self.account_mapper
        )
        return builder.gerar()
    
    def _aplicar_formatacao(self, ws, num_cols: int, num_rows: int, coluna_codigo_texto: Optional[int] = None, colunas_texto: Optional[List[int]] = None, mapa_tipo_conta: Optional[Dict[str, str]] = None):
        """
        Aplica formatação básica à planilha.
        
        Args:
            ws: Worksheet do openpyxl
            num_cols: Número de colunas
            num_rows: Número de linhas
            coluna_codigo_texto: Coluna que deve ser formatada como texto (sem formatação numérica) - deprecated, use colunas_texto
            colunas_texto: Lista de colunas (1-indexed) que devem ser formatadas como texto
            mapa_tipo_conta: Dicionário mapeando código da conta (str) -> TIPO_CTA ("S" ou "A")
        """
        # Compatibilidade: se coluna_codigo_texto for fornecido, adiciona à lista
        if colunas_texto is None:
            colunas_texto = []
        if coluna_codigo_texto is not None and coluna_codigo_texto not in colunas_texto:
            colunas_texto.append(coluna_codigo_texto)
        # Estilos
        font_titulo = Font(bold=True, size=12)
        font_subtitulo = Font(bold=True, size=11)
        font_normal = Font(size=10)
        align_right = Alignment(horizontal="right")
        align_left = Alignment(horizontal="left")
        
        # Formato de moeda brasileiro
        moeda_format = '#,##0.00'  # Será aplicado via número
        
        # Borda
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplica formatação
        for row in range(1, min(num_rows + 1, ws.max_row + 1)):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                
                # Formatação de texto
                if row == 1:  # Cabeçalho
                    cell.font = font_titulo
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.value and isinstance(cell.value, str):
                    valor_str = str(cell.value)
                    deve_estar_negrito = False
                    
                    # Só verifica formatação em negrito para a primeira coluna (nome da conta/item)
                    # Outras colunas (valores numéricos) não precisam dessa verificação
                    if col == 1:
                        # Verifica se é total ou subtotal (palavras-chave específicas)
                        palavras_chave_totais = ["TOTAL", "ATIVO", "PASSIVO", "PATRIMÔNIO", "RECEITAS", "DESPESAS", "CUSTOS", "RESULTADO"]
                        contem_palavra_chave = any(keyword in valor_str.upper() for keyword in palavras_chave_totais)
                        
                        # Tenta extrair código da conta entre parênteses
                        match = re.search(r'\(([^)]+)\)', valor_str)
                        codigo_conta = None
                        if match:
                            codigo_conta = match.group(1).strip()
                        
                        if contem_palavra_chave:
                            # Contém palavras-chave de totais
                            if codigo_conta:
                                # Tem código entre parênteses - verifica TIPO_CTA
                                if mapa_tipo_conta and codigo_conta in mapa_tipo_conta:
                                    tipo_cta = mapa_tipo_conta[codigo_conta]
                                    deve_estar_negrito = (tipo_cta == "S")  # Sintética = negrito
                                else:
                                    # Não encontrou no mapa - assume total/subtotal (negrito)
                                    deve_estar_negrito = True
                            else:
                                # Não tem código - é total/subtotal (negrito)
                                deve_estar_negrito = True
                        else:
                            # Não contém palavras-chave de totais
                            if codigo_conta:
                                # Tem código entre parênteses - verifica TIPO_CTA
                                if mapa_tipo_conta and codigo_conta in mapa_tipo_conta:
                                    tipo_cta = mapa_tipo_conta[codigo_conta]
                                    deve_estar_negrito = (tipo_cta == "S")  # Sintética = negrito
                                else:
                                    # Não encontrou no mapa - assume normal
                                    deve_estar_negrito = False
                            else:
                                # Não tem código e não é total - assume normal
                                deve_estar_negrito = False
                    
                    # Aplica formatação
                    if deve_estar_negrito:
                        cell.font = font_subtitulo
                    else:
                        cell.font = font_normal
                    cell.alignment = align_left
                
                # Formatação numérica (exceto para colunas de texto especificadas)
                if col in colunas_texto and cell.value is not None:
                    # Colunas de texto: mantém como texto, sem formatação numérica
                    cell.number_format = '@'  # Formato texto
                    cell.alignment = align_left
                elif col not in colunas_texto and isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                
                # Borda
                cell.border = thin_border
        
        # Autoajusta largura das colunas
        for col in range(1, num_cols + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def exportar_excel(self, outdir: Path, nome_arquivo: Optional[str] = None) -> Path:
        """
        Exporta dados contábeis para arquivo Excel.
        
        Args:
            outdir: Diretório de saída
            nome_arquivo: Nome do arquivo (opcional, será gerado automaticamente se None)
            
        Returns:
            Caminho do arquivo Excel gerado
        """
        # Cria diretório se não existir
        outdir = Path(outdir)
        outdir.mkdir(parents=True, exist_ok=True)
        
        # Nome do arquivo
        if nome_arquivo is None:
            nome_arquivo = f"contabilidade_{self.empresa}_{self.inicio}_{self.fim}.xlsx"
        
        excel_path = outdir / nome_arquivo
        
        # Busca dados
        if self.df_pc is None:
            self.buscar_plano_contas_com_saldos()
        
        if self.df_movimentacoes is None:
            self.buscar_movimentacao_periodo()
        
        if self.df_lancamentos is None:
            self.buscar_lancamentos_periodo()
        
        # Cria workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove planilha padrão
        
        # Aba 1: Plano de Contas
        if self.df_pc is not None and not self.df_pc.empty:
            ws_pc = wb.create_sheet("Plano de Contas")
            df_pc_export = self.df_pc[["CODI_CTA", "NOME_CTA", "CLAS_CTA", "TIPO_CTA", "SITUACAO_CTA", "BC_ACCOUNT"]].copy()
            df_pc_export = df_pc_export.sort_values("CLAS_CTA")
            
            # Cabeçalho
            headers = ["Código", "Nome", "Classificação", "Tipo", "Situação", "Classificação Beancount"]
            ws_pc.append(headers)
            
            # Dados
            for _, row in df_pc_export.iterrows():
                # Código como texto para evitar formatação numérica
                codigo = str(row["CODI_CTA"]) if pd.notna(row["CODI_CTA"]) else ""
                ws_pc.append([
                    codigo,
                    row["NOME_CTA"],
                    row["CLAS_CTA"],
                    row["TIPO_CTA"],
                    row["SITUACAO_CTA"],
                    row["BC_ACCOUNT"]
                ])
            
            # Cria mapa de TIPO_CTA para formatação
            mapa_tipo_conta = self._criar_mapa_tipo_conta()
            self._aplicar_formatacao(ws_pc, len(headers), len(df_pc_export) + 1, coluna_codigo_texto=1, mapa_tipo_conta=mapa_tipo_conta)
        
        # Aba 2: Balanço Patrimonial
        df_bp = self.gerar_balanco_patrimonial()
        if not df_bp.empty:
            ws_bp = wb.create_sheet("Balanço Patrimonial")
            headers = ["Conta/Categoria", "Saldo"]
            ws_bp.append(headers)
            
            for _, row in df_bp.iterrows():
                ws_bp.append([row["Conta/Categoria"], row["Saldo"]])
            
            # Cria mapa de TIPO_CTA para formatação
            mapa_tipo_conta = self._criar_mapa_tipo_conta()
            self._aplicar_formatacao(ws_bp, len(headers), len(df_bp) + 1, mapa_tipo_conta=mapa_tipo_conta)
        
        # Aba 3: DRE
        df_dre = self.gerar_dre()
        if not df_dre.empty:
            ws_dre = wb.create_sheet("DRE")
            
            # Cabeçalhos dinâmicos baseados nas colunas do DataFrame
            headers = df_dre.columns.tolist()
            ws_dre.append(headers)
            
            # Dados
            for _, row in df_dre.iterrows():
                linha = []
                for col in headers:
                    valor = row.get(col)
                    linha.append(valor)
                ws_dre.append(linha)
            
            # Identifica colunas numéricas (todas exceto "Item")
            colunas_numericas = [i + 1 for i, col in enumerate(headers) if col != "Item"]
            
            # Cria mapa de TIPO_CTA para formatação
            mapa_tipo_conta = self._criar_mapa_tipo_conta()
            # Aplica formatação
            self._aplicar_formatacao(ws_dre, len(headers), len(df_dre) + 1, colunas_texto=[1], mapa_tipo_conta=mapa_tipo_conta)
        
        # Aba 4: Movimentação do Período
        if self.df_lancamentos is not None and not self.df_lancamentos.empty:
            ws_mov = wb.create_sheet("Movimentação do Período")
            
            # Usa PeriodMovementsBuilder para gerar o extrato
            builder = PeriodMovementsBuilder(self.df_lancamentos, self.account_mapper)
            df_mov_export = builder.gerar()
            
            if not df_mov_export.empty:
                # Cabeçalhos: Data, Código Débito, Conta Débito, Código Crédito, Conta Crédito, Histórico, Documento, Lote, Valor
                headers = ["Data", "Código Débito", "Conta Débito", "Código Crédito", "Conta Crédito", "Histórico", "Documento", "Lote", "Valor"]
                ws_mov.append(headers)
                
                # Função helper para converter números inteiros corretamente
                def formatar_numero_texto(valor):
                    """Converte valor numérico para string, removendo .0 de inteiros."""
                    if pd.isna(valor):
                        return ""
                    # Se for numérico e inteiro, converte para int primeiro
                    if isinstance(valor, (int, float)):
                        if float(valor).is_integer():
                            return str(int(valor))
                        return str(valor)
                    str_val = str(valor).strip()
                    if str_val in ["nan", "None", ""]:
                        return ""
                    return str_val
                
                for _, row in df_mov_export.iterrows():
                    # Converte codi_lote e ndoc_lan para string (formato texto)
                    codi_lote_val = row.get("codi_lote", "")
                    codi_lote_str = formatar_numero_texto(codi_lote_val)
                    if codi_lote_str == "0":
                        codi_lote_str = ""
                    
                    ndoc_lan_val = row.get("ndoc_lan", "")
                    ndoc_lan_str = formatar_numero_texto(ndoc_lan_val)
                    
                    ws_mov.append([
                        row.get("data_lan", ""),
                        row.get("Código Débito", ""),
                        row.get("Conta Débito", ""),
                        row.get("Código Crédito", ""),
                        row.get("Conta Crédito", ""),
                        row.get("chis_lan", ""),
                        ndoc_lan_str,
                        codi_lote_str,
                        row.get("vlor_lan", 0)
                    ])
                
                # Colunas de texto: 2 (Código Débito), 4 (Código Crédito), 7 (Documento), 8 (Lote)
                colunas_texto = [2, 4, 7, 8]
                # Cria mapa de TIPO_CTA para formatação
                mapa_tipo_conta = self._criar_mapa_tipo_conta()
                self._aplicar_formatacao(ws_mov, len(headers), len(df_mov_export) + 1, colunas_texto=colunas_texto, mapa_tipo_conta=mapa_tipo_conta)
        
        # Aba 5: Balancete
        df_balancete = self.gerar_balancete()
        if not df_balancete.empty:
            ws_balancete = wb.create_sheet("Balancete")
            headers = ["Código", "Nome", "Classificação", "Saldo Inicial", "Total Débitos", "Total Créditos", "Saldo Final"]
            ws_balancete.append(headers)
            
            for _, row in df_balancete.iterrows():
                # Código como texto para evitar formatação numérica
                codigo = str(row["Código"]) if pd.notna(row["Código"]) else ""
                ws_balancete.append([
                    codigo,
                    row["Nome"],
                    row["Classificação"],
                    row["Saldo Inicial"],
                    row["Total Débitos"],
                    row["Total Créditos"],
                    row["Saldo Final"]
                ])
            
            # Cria mapa de TIPO_CTA para formatação
            mapa_tipo_conta = self._criar_mapa_tipo_conta()
            self._aplicar_formatacao(ws_balancete, len(headers), len(df_balancete) + 1, coluna_codigo_texto=1, mapa_tipo_conta=mapa_tipo_conta)
        
        # Salva arquivo
        wb.save(excel_path)
        
        return excel_path

